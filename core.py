# core.py
import logging
import traceback
import time

from services.deepseek_client import get_client  # ✅ 用service层

# ====== 初始化 client（全局唯一）======
client = get_client()


logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)


# ====== AI调用 ======
def call_ai(prompt):
    """
    调用AI生成测试用例
    """
    client = get_client()   # ✅ 统一从 service 层获取

    try:
        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {
                    "role": "system",
                    "content": "你是一名资深测试架构师，必须严格按照JSON格式输出"
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0.3,   # 降低随机性（更稳定）
            max_tokens=1000
        )

        result = response.choices[0].message.content

        # ✅ 基本校验（防AI乱输出）
        if not result or len(result.strip()) == 0:
            raise ValueError("AI返回内容为空")

        return result

    except Exception as e:
        # ✅ 完整日志（工程必备）
        logging.error("调用AI失败")
        logging.error(f"错误类型: {type(e).__name__}")
        logging.error(f"错误信息: {str(e)}")
        logging.error(traceback.format_exc())

        # ✅ 简单重试一次（抗网络波动）
        try:
            time.sleep(2)

            response = client.chat.completions.create(
                model="deepseek-chat",
                messages=[
                    {"role": "system", "content": "你是一名资深测试架构师，必须严格按照JSON格式输出"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=1000
            )

            return response.choices[0].message.content

        except Exception:
            logging.error("重试失败")

            # ❗返回给GUI的友好提示
            return "❌ AI请求失败（可能是网络/API限流），请稍后重试"


def generate_case(user_input, type_choice="3", callback=None):
    logging.info("开始生成测试用例")

    try:
        # ✅ 1. 类型选择
        if type_choice == "1":
            type_desc = "只生成功能测试用例"
        elif type_choice == "2":
            type_desc = "只生成异常测试用例"
        else:
            type_desc = "生成全部测试用例"

        # ✅ 2. 构造 prompt
        prompt = f"""
产品需求：
{user_input}

要求：
{type_desc}

请严格按照JSON格式输出测试用例列表：
[
  {{
    "case_id": "TC001",
    "title": "测试点",
    "steps": "步骤",
    "expected": "预期结果"
  }}
]
要求：
1. 必须是合法JSON（不能有多余解释）
2. 至少生成5条测试用例
"""

        # ✅ 3. UI提示
        if callback:
            callback("正在请求AI...")

        # ✅ 4. 调用AI
        result = call_ai(prompt)

        # ✅ 5. UI更新
        if callback:
            callback("生成完成")

        logging.info("测试用例生成成功")

        # ✅ 6. 成功返回
        return {
            "success": True,
            "message": "生成成功",
            "data": result
        }

    except Exception as e:
        logging.error("生成失败", exc_info=True)

        if callback:
            callback("生成失败")

        # ❗核心：兜底返回
        return {
            "success": False,
            "message": "生成失败，请稍后重试",
            "data": None
        }


# ====== 导出Excel（自动列宽+自动换行） ======
def export_to_excel(data_text, file_path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        import json
        import re

        # ====== 清洗函数 ======
        def clean_text(text):
            text = text.replace("<br>", "\n")
            text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
            text = re.sub(r"\*(.*?)\*", r"\1", text)
            return text.strip()

        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        headers = ["用例ID", "标题", "步骤", "预期结果"]
        ws.append(headers)

        # ====== 表头样式 ======
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # =========================
        # 🚀 ① 尝试 JSON 解析
        # =========================
        json_match = re.search(r"\[.*\]", data_text, re.S)

        if json_match:
            data = json.loads(json_match.group())

            for item in data:
                ws.append([
                    item.get("case_id", ""),
                    item.get("title", ""),
                    item.get("steps", ""),
                    item.get("expected", "")
                ])

        else:
            # =========================
            # 🚀 ② 兼容 Markdown 表格
            # =========================
            lines = data_text.split("\n")
            table_lines = [line for line in lines if "|" in line]

            if len(table_lines) < 2:
                return "❌ 未识别到表格或JSON"

            headers_md = [clean_text(h) for h in table_lines[0].split("|") if h.strip()]

            for line in table_lines[2:]:
                cols = [clean_text(c) for c in line.split("|") if c.strip()]
                if len(cols) == len(headers_md):
                    ws.append(cols)

        # ====== 自动换行 ======
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # ====== 自适应列宽 ======
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_length + 5, 60)

        wb.save(file_path)

        return f"✅ 导出成功：{file_path}"

    except Exception as e:
        return f"❌ 导出失败：{str(e)}"